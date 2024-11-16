
# print("Ajide Shedrack")
# print("Number")
# print("0----")
# print(" |||| ")

price = (12 * 5)
# print(price)

name = "John Smith"
age = 20
registered = "May 12 2024"

# the_name = input("What is your name? ")
# print("Hi " + the_name)

# the_name = input("What is your name? ")
# the_color = input("What is your favorite? ")
#
# print(the_name + " like " + the_color)

# pounds_weight = input("What is your weight in pounds? ")
# kg = float(pounds_weight) * 0.45
# print(kg)


course = "python for beginner"
# print(course[2])
# print(course[0:3])
# print(course[0:])
# print(course[3:])
# print(course[:8])
# print(course[:]) # by default clone a string.


first = "Shabs"
last = "Law"
msg = f'{first} [{last}] is a coder'
# print(msg)
# print(first.replace("s", "slaw"))
# print("h" in last)


# print(2 + 3)
# print(2 * 3)
# print(2 - 3)
# print(3 / 2)
# print(3 % 2)
# print(2 // 3)
# print(2 ** 3)


x = 10 + 2 * 3
# print(x) # 16


is_hot = False
is_cold = True

# if is_hot:
#     print("It's a hot day")
#     print("drink a lot of water")
# elif is_cold:
#     print("It's a cold day")
#     print("Use worm blanket")
# else:
#     print("is a normal day")
#
# print("Enjoy your day.")

good_credit = False
bad_credit = False
house_cost = 1000000
down_payment = 0

if good_credit:
     down_payment = house_cost / 100 * 10
     print(down_payment)
elif bad_credit:
    down_payment = house_cost / 100 * 20
    print(down_payment)
else:
    print(down_payment)


# Logical Operator

# AND: both condition should be true
# OR: one of the condition should be true
# NOT: with a false condition assign to the NOT operation it is true.


# Comperising Operator

# 1 > 2  : 1 is greater than 2
# 1 < 2  : 1 is less than 2

p_name = "shabs"

if len(p_name) < 3:
    print("Name most be at list 3 Characters long")
elif len(p_name) > 50:
    print("Name can be maximum of 50 Characters")
else:
    print("Name looks good.")



# weight = input("weight: ")
# l_or_k = input("(L)bs or (K)g: ")
# if l_or_k.upper() == "L":
#     answer = int(weight) // 2.205
#     print(f'You are {answer} kilos.')
# elif l_or_k.upper() == "K":
#     answer = int(weight) * 2.205
#     print(f'You are {round(answer)} Pounds.')


# While Loop
i = 1

# while i <= 5:
#     print(i)
#     i = i + 1
# print("Done")


# while i <= 5:
#     print("Shabs" * i)
#     i = i + 1
# print("Done")


# GUESSING GAME

# guess_num = 9
# count_num = 1
# count_limit = 3
#
# while count_num <= count_limit:
#     guess = int(input("Guess a Number: "))
#     count_num += 1
#     if guess == guess_num:
#         print("You Win!")
#         break
# else:
#     print("Sorry, you Lost!")



# #     CAR GAME
# help = "help"
# start = "start"
# stop = "stop"
# quit = "quit"
# count_num = 1
# count_limit = 0
# started = False
#
#
# while count_num >= count_limit:
#     guess = input("> ").lower()
#     count_num += 1
#     if guess == help:
#         print('''
# start - to start the car
# stop - to stop the car
# quit - to quit the car
#         ''')
#     elif guess == start:
#         if started:
#             print("Car already Started!")
#         else:
#             started = True
#             print("Car Started...Ready to go!")
#     elif guess == stop:
#         if not started:
#             print("Car is already Stop!")
#         else:
#             started = False
#             print("Car Stopped.")
#     elif guess == quit:
#         break
#     else:
#         print("I don't understand that...type 'Help'")




# FOR LOOP

# for item in "Shabs":
#     print(item)

# for item in range(20):
# for item in range(5, 20):
# for item in range(5, 20, 3):
#     print(item)

prices = [10, 20, 30]
total = 0

for price in prices:
    total += price
# print(total)



# NESTED LOOPs

# for x in range(4):
#     for y in range(3):
#         print(f"({x}), ({y})")


price_lists = [5, 2, 5, 2, 2]
for item in price_lists:
    output = ''
    for x in range(item):
        output += "x"
    # print(output)




# LISTS

numbers = [5, 2, 4, 8, 10, 2]
max = numbers[0]

for number in numbers:
    if number > max:
        max = number

# print(max)



# LIST METHOD

# numbers.append(20)  # add number to the end
# numbers.insert(2, 10)    # add an item after the index 2 of numbers list
# numbers.remove(10)   # to remove item by the name of the item
# numbers.clear()    # to remove all the items in the list
# numbers.pop()      # to remove the last item on the list
# numbers.index(5)   # to check what index a number exist
# 5 in numbers       # to check if a number exist in a list
# numbers.count(2)   # to count the amount an exact number we have in a list
# numbers.sort()     # to sort how list from  1 - 100 in an assending other
# numbers.reverse()  # to reverse a list by desending other
# numbers.copy()     # to copy a list




# REMOVE DUPLICATE

lists = [2, 4, 3, 2, 2, 4, 6, 7, 4]
unique = []

for list in lists:
    if list not in unique:
        unique.append(list)

# print(unique)




# TUPLE

tuple_lists = (2,3,4,5)



# UNPACKING

coordinate = (2,3,4)
x = coordinate[0]
y = coordinate[1]
z = coordinate[2]

# or

x, y, z = coordinate



# DICTIONARIES

person = {
    "name": "John Don",
    "age": 30,
    "is_verified": True,
}

person["name"] = "Shabs Planta"
person["DOB"] = "Jan 11 2023"
person.get("birthdate")
person.get("birthdate", "Jan - 10 - 1994")
# print(person)


# phone = input("Phone: ")
# number_mapping = {
#     "1": "one",
#     "2": "two",
#     "3": "three",
#     "4": "four"
# }
# output_num = ""
#
# for ch in phone:
#     output_num += number_mapping.get(ch, "!") + " "
# print(output_num)


# phone = input("> ")
# num_mapping = {
#     "1": "one",
#     "2": "two",
#     "3": "three",
#     "4": "four"
# }
# out_put = ""
#
# for ch in phone:
#     out_put += num_mapping.get(ch, ch) + " "
# print(out_put)



#        FUNCTIONS

def greetings():
    print("Hello There!")
    print("Thank You For Coming.")


# print("Start")
# greetings()
# print("Finish")




def replay(tell):
    num_mapping = {
        "1": "one",
        "2": "two",
        "3": "three",
        "4": "four"
    }
    out_put = ""

    for ch in tell:
        out_put += num_mapping.get(ch, ch) + " "
    return out_put

# phone = input("> ")
# print(replay(phone))



#    EXCEPTIONS -  TRY & EXCEPT ERROR
# try:
#     age = int(input("Age: "))
#     income = 20000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print("Input value can't be 0.")
# except ValueError:
#     print("Invalued input")



#   CLASS
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point()
point1.x = 10
point1.y = 20

# print(point1.x)
# print(point1.draw())



#   CLASS  with Creating a CONSTRUCTOR
class Point:
    def __init__(self, x , y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point(10, 20)
point1.x = 15

# print(point1.x)
# print(point1.draw())


class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        return f"Hello everyone, I'm {self.name}"


id = Person("Shabs")
# print(id.name)
# print(id.talk())



# INHERITANCE

class Mammal:
    def work(self):
        print("Working")

class Dog(Mammal):
    pass

class Cat(Mammal):
    pass


dog1 = Dog()
# dog1.work()



# IMPORTATION
import  converters
from converters import kg_to_lbs

# print(converters.kg_to_lbs(72))
# print(converters.lbs_to_kg(72))
# print(kg_to_lbs(80))


import  ecommerce.coverterss
from ecommerce import coverterss
from ecommerce.coverterss import international_data

# international_data()
# coverterss.international_data()
# ecommerce.coverterss.international_data()



import random

for i in range(3):
    break
    # print(random.random())
    # print(random.randint(10, 20))

members = ["John", "Mary", "Tope", "Ayo"]
# print(random.choice(members))




# DICE GAME
class Dice:
    def roll(self):
        return (f"{random.randint(1, 6)}", f"{random.randint(1, 6)}")

dice1 = Dice()
# print(dice1.roll())



# PATHs - PATH LIBIRIES
from pathlib import Path

# Absolute path
# c:/user/local/myfil

# Relative path
path = Path()
created_path = Path("ecommerce")
new_path = Path("email")

# print(path.glob('*.*'))
# print(created_path.exists())
# print(new_path.mkdir())
# print(new_path.rmdir())

# for i in path.glob('*.py'):
for i in path.glob('*'):
    break
    # print(i)







# WORKING WITH EXCEL ON PYTHON

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wd = xl.load_workbook("transactions.xlsx")
sheet = wd["Sheet1"]
# cell = sheet["a1"]
# cell = sheet.cell(1, 1)

# print(cell.value)
# print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wd.save('transactions2.xlsx')



# for a clean code.
def process_workbook(workbook, sheetbook):
    wd_in = xl.load_workbook(workbook)
    sheet_in = wd[sheetbook]

    for row_in in range(2, sheet.max_row + 1):
        cell_in = sheet.cell(row, 3)
        corrected_price_in = cell_in.value * 0.9
        corrected_price_in_cell = sheet.cell(row, 4)
        corrected_price_in_cell.value = corrected_price_in

    values_in = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart_in = BarChart()
    chart_in.add_data(values_in)
    sheet_in.add_chart(chart_in, 'e2')

    wd_in.save(workbook)